"""
Скрипт извлекает данные из xlsx-файлов и сохраняет в txt.
Запуск: python extract_finances.py
"""
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Устанавливаю openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

BASE = r"C:\Users\user\Desktop\MyBusiness\inbox"
OUT  = r"C:\Users\user\Desktop\MyBusiness\finances_export.txt"

files = [
    "Итоги за 2025 год.xlsx",
    "finmodel_pervouralsk_fix_editable_v7 (1).xlsx",
]

lines = []

for fname in files:
    fpath = os.path.join(BASE, fname)
    if not os.path.exists(fpath):
        lines.append(f"[НЕ НАЙДЕН] {fname}\n")
        continue

    lines.append(f"\n{'='*70}")
    lines.append(f"ФАЙЛ: {fname}")
    lines.append(f"{'='*70}")

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"\n--- Лист: {sheet_name} ---")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                if any(cell is not None for cell in row):
                    row_str = "\t".join(
                        str(c) if c is not None else "" for c in row
                    )
                    lines.append(row_str)
    except Exception as e:
        lines.append(f"[ОШИБКА чтения] {e}")

result = "\n".join(lines)
with open(OUT, "w", encoding="utf-8") as f:
    f.write(result)

print(f"Готово! Файл сохранён: {OUT}")
print("Скопируй содержимое finances_export.txt и вставь в чат с Claude.")
